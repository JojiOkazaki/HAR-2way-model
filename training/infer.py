# infer.py
import os
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, List, Tuple

import yaml
import torch
import torch.nn.functional as F
from torch.utils.data import Dataset, DataLoader

import numpy as np
from openpyxl import Workbook
import matplotlib.pyplot as plt

# --- project modules (uploaded files) ---
from training.modules.models import CNN
from training.modules.models import STGCN
from training.modules.models import MLP
from training.modules.models import TransformerEncoder
from training.modules.networks import ImageBranch
from training.modules.networks import SkeletonBranch
from training.modules.networks import FullModel
from training.modules.utils import build_coco17_adj
from config_base import *

try:
    from config_local import *
except ImportError:
    pass

# -----------------------------
# config helpers
# -----------------------------
def load_yaml(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


# -----------------------------
# dataset with class_id
# -----------------------------
class SyncedDatasetWithClassId(Dataset):
    """
    file_list: each line "ClassName/xxx.pt <class_id>"
    returns:
      frames:   (P, T, C, H, W)
      keypoints:(P, T, J, C_kp)
      scores:   (P, T, J)
      labels:   (P, D)  # stored in image pt
      class_id: int
    """
    def __init__(self, img_torch_path: Path, skel_torch_path: Path, file_list: Path):
        self.img_torch_path = img_torch_path
        self.skel_torch_path = skel_torch_path

        self.samples: List[Path] = []
        self.class_ids: List[int] = []

        with file_list.open("r", encoding="utf-8") as f:
            for line_no, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue
                parts = line.split()
                if len(parts) != 2:
                    raise ValueError(f"{file_list}:{line_no}: expected 2 columns, got {len(parts)}: {line!r}")
                relpath, id_str = parts
                img_path = self.img_torch_path / relpath
                if not img_path.exists():
                    raise FileNotFoundError(f"image pt not found: {img_path}")
                self.samples.append(img_path)
                self.class_ids.append(int(id_str))

    def __len__(self) -> int:
        return len(self.samples)

    def __getitem__(self, idx: int):
        img_path = self.samples[idx]
        class_id = self.class_ids[idx]

        labelname = img_path.parent.name
        filename = img_path.name
        skel_path = self.skel_torch_path / labelname / filename
        if not skel_path.exists():
            raise FileNotFoundError(f"skeleton pt not found: {skel_path}")

        # NOTE: weights_only=True requires newer PyTorch; if your env errors, remove it.
        img_data = torch.load(img_path, weights_only=True)
        skel_data = torch.load(skel_path, weights_only=True)

        frames = img_data["images"].float()        # (P,T,C,H,W)
        labels = img_data["labels"].float()        # (P,D)

        keypoints = skel_data["skeletons"].float() # (P,T,J,C_kp)
        scores = skel_data["scores"].float()       # (P,T,J)

        return frames, keypoints, scores, labels, class_id


# -----------------------------
# model builder (same as training)
# -----------------------------
def create_model(params: dict, device: torch.device) -> torch.nn.Module:
    params = dict(params)  # shallow copy
    params["skel"]["stgcn"] = dict(params["skel"]["stgcn"])
    params["skel"]["stgcn"]["adj"] = build_coco17_adj(device=device)

    model = FullModel(
        ImageBranch(
            CNN(**params["img"]["cnn"]),
            TransformerEncoder(**params["img"]["transformer"]),
        ),
        SkeletonBranch(
            STGCN(**params["skel"]["stgcn"])
        ),
        MLP(**params["mlp"]),
    )
    return model


# -----------------------------
# sentence embeddings (class candidates)
# -----------------------------
def load_class_texts_from_csv(csv_path: Path) -> Tuple[List[int], List[str]]:
    import pandas as pd
    df = pd.read_csv(csv_path)
    if "label_id" not in df.columns or "sentence" not in df.columns:
        raise ValueError(f"CSV must have columns: label_id, sentence. got: {list(df.columns)}")
    df = df.sort_values("label_id")
    label_ids = df["label_id"].astype(int).tolist()
    sentences = df["sentence"].astype(str).tolist()

    # sanity: expect 0..C-1 contiguous
    if label_ids and (min(label_ids) != 0 or max(label_ids) != len(label_ids) - 1):
        raise ValueError(f"label_id must be contiguous 0..C-1. got min={min(label_ids)}, max={max(label_ids)}, C={len(label_ids)}")

    return label_ids, sentences


def encode_class_texts(sentences: List[str], model_name: str, device: str) -> torch.Tensor:
    """
    returns: (C, D) float32 on CPU
    """
    try:
        from sentence_transformers import SentenceTransformer
    except Exception as e:
        raise RuntimeError(
            "sentence-transformers is required for inference candidates. "
            "Install: pip install sentence-transformers"
        ) from e

    st = SentenceTransformer(model_name, device=device)
    emb = st.encode(
        sentences,
        convert_to_numpy=True,
        normalize_embeddings=True,  # cosine用に正規化して保存
        show_progress_bar=True,
        batch_size=64,
    )
    emb_t = torch.from_numpy(emb).float().cpu()  # (C,D)
    return emb_t


# -----------------------------
# metrics
# -----------------------------
@dataclass
class MetricState:
    recall_hits: Dict[int, int]            # k -> hits
    total: int
    confusion_by_k: Dict[int, np.ndarray]  # k -> (C,C) float64


def update_metrics(
    state: MetricState,
    sim_logits: torch.Tensor,     # (N,C)
    true_ids: torch.Tensor,       # (N,)
    ks: List[int],
):
    # recall@k + confusion@k
    t = true_ids.detach().cpu().numpy().astype(np.int64)

    for k in ks:
        k_eff = min(k, sim_logits.size(1))
        topk = sim_logits.topk(k_eff, dim=1).indices  # (N,k_eff)

        # recall@k
        hit = (topk == true_ids.unsqueeze(1)).any(dim=1)
        state.recall_hits[k] += int(hit.sum().item())

        # confusion@k: 1サンプルを top-k へ 1/k ずつ配る
        p = topk.detach().cpu().numpy().astype(np.int64)  # (N,k_eff)
        w = 1.0 / float(k_eff)
        cm = state.confusion_by_k[k]

        for ti, row in zip(t, p):
            for pj in row:
                cm[ti, pj] += w

    state.total += int(true_ids.numel())


def save_confusion_excel(
    out_path: Path,
    confusion: np.ndarray,
    sentences: List[str],
):
    C = confusion.shape[0]
    wb = Workbook()
    ws = wb.active
    ws.title = "confusion"

    # header
    ws.cell(row=1, column=1, value="true\\pred")
    for j in range(C):
        ws.cell(row=1, column=2 + j, value=j)

    for i in range(C):
        ws.cell(row=2 + i, column=1, value=i)
        for j in range(C):
            ws.cell(row=2 + i, column=2 + j, value=round(float(confusion[i, j]), 6))


    # mapping sheet
    ws2 = wb.create_sheet("label_text")
    ws2.cell(row=1, column=1, value="label_id")
    ws2.cell(row=1, column=2, value="sentence")
    for i, s in enumerate(sentences):
        ws2.cell(row=2 + i, column=1, value=i)
        ws2.cell(row=2 + i, column=2, value=s)

    wb.save(out_path)


def save_confusion_png(
    out_path: Path,
    confusion: np.ndarray,
    normalize: str = "true",  # "none" / "true" / "pred" / "all"
):
    cm = confusion.astype(np.float32)

    if normalize == "true":
        denom = cm.sum(axis=1, keepdims=True)
        denom[denom == 0] = 1.0
        cm = cm / denom
    elif normalize == "pred":
        denom = cm.sum(axis=0, keepdims=True)
        denom[denom == 0] = 1.0
        cm = cm / denom
    elif normalize == "all":
        denom = cm.sum()
        denom = denom if denom != 0 else 1.0
        cm = cm / denom
    elif normalize == "none":
        pass
    else:
        raise ValueError(f"unknown normalize={normalize}")

    plt.figure(figsize=(12, 10))
    plt.imshow(cm, aspect="auto")
    plt.title(f"Confusion Matrix (normalize={normalize})")
    plt.xlabel("Predicted")
    plt.ylabel("True")
    plt.colorbar()
    plt.tight_layout()
    plt.savefig(out_path, dpi=200)
    plt.close()


def main():
    infer_cfg_path = Path("training/configs/infer.yaml")
    infer_cfg = load_yaml(infer_cfg_path)

    artifact_path = infer_cfg["artifact_path"]
    run_dir = (ARTIFACT_ROOT / artifact_path).resolve()
    if not run_dir.exists():
        raise FileNotFoundError(f"run_dir not found: {run_dir}")

    timestamp = run_dir.name
    train_cfg_path = run_dir / f"{timestamp}_config.yaml"
    train_cfg = load_yaml(train_cfg_path)

    # resolve best model path (training rule)
    best_model_name = f'{timestamp}_{train_cfg["logging"]["best_model_name"]}'
    ckpt_path = run_dir / best_model_name
    if not ckpt_path.exists():
        raise FileNotFoundError(f"checkpoint not found: {ckpt_path}")

    # infer settings
    infer_block = infer_cfg.get("infer", {})
    test_file_list = DATASET_ROOT / infer_block["test_file_list"]
    recall_ks = [int(k) for k in infer_block.get("recall_ks", [1, 5, 10])]
    targets = infer_block.get("targets", ["full"])
    out_dir = run_dir / infer_block.get("out_dir", "infer")
    ensure_dir(out_dir)

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    st_device = "cuda" if torch.cuda.is_available() else "cpu"

    # dataset paths from training config
    img_pt_dir = DATASET_ROOT / train_cfg["data"]["img_pt_dir"]
    skel_pt_dir = DATASET_ROOT / train_cfg["data"]["skel_pt_dir"]
    num_workers = int(train_cfg["data"].get("num_workers", 4))
    batch_size = int(train_cfg["training"]["batch_size"]) if "training" in train_cfg else 8

    # build dataloader
    ds = SyncedDatasetWithClassId(img_pt_dir, skel_pt_dir, test_file_list)
    dl = DataLoader(
        ds,
        batch_size=batch_size,
        shuffle=False,
        num_workers=num_workers,
        pin_memory=True,
        persistent_workers=(num_workers > 0),
        drop_last=False,
    )

    # build model
    model_params = train_cfg["model"]["architecture"]
    model = create_model(model_params, device=device).to(device)
    state_dict = torch.load(ckpt_path, map_location=device, weights_only=True)
    model.load_state_dict(state_dict, strict=True)
    model.eval()

    # class candidate embeddings
    csv_path = DATASET_ROOT / "ucf101_sentences_jp.csv"
    label_ids, sentences = load_class_texts_from_csv(csv_path)
    class_emb = encode_class_texts(
        sentences=sentences,
        model_name=infer_block["sentence_transformer"],
        device=st_device,
    )  # (C,D) normalized, CPU float
    C, D = class_emb.shape

    class_emb = class_emb.to(device)
    class_emb = F.normalize(class_emb, dim=-1)  # 念のため

    # metrics per target
    metrics: Dict[str, MetricState] = {}
    for t in targets:
        metrics[t] = MetricState(
            recall_hits={k: 0 for k in recall_ks},
            total=0,
            confusion_by_k={k: np.zeros((C, C), dtype=np.float64) for k in recall_ks},
        )


    # inference loop
    with torch.inference_mode():
        for frames, keypoints, scores, _labels, class_id in dl:
            # frames:   (B,P,T,C,H,W)
            # keypoints:(B,P,T,J,C_kp)
            # scores:   (B,P,T,J)
            B, P = scores.shape[:2]

            frames = frames.to(device, non_blocking=True)
            keypoints = keypoints.to(device, non_blocking=True)
            scores = scores.to(device, non_blocking=True)
            class_id = torch.as_tensor(class_id, device=device, dtype=torch.long)  # (B,)
            class_id = class_id - 1

            # valid persons mask (same logic as training)
            # scores.max(-1) -> (B,P,T)
            person_has_keypoint = (scores.max(dim=-1).values > 0)
            # any over T -> (B,P)
            person_valid = person_has_keypoint.any(dim=-1)
            valid_flat = person_valid.reshape(-1)  # (B*P,)

            idx = valid_flat.nonzero(as_tuple=False).squeeze(1)
            if idx.numel() == 0:
                continue

            # flatten (B*P, ...)
            frames_f = frames.reshape(B * P, *frames.shape[2:])[idx]
            keypoints_f = keypoints.reshape(B * P, *keypoints.shape[2:])[idx]
            scores_f = scores.reshape(B * P, *scores.shape[2:])[idx]

            # repeat class_id per person
            class_id_f = class_id.repeat_interleave(P, dim=0)[idx]  # (N,)

            # forward: returns (full, img, skel)
            full_emb, img_emb, skel_emb = model(frames_f, keypoints_f, scores_f)
            outs: Dict[str, torch.Tensor] = {
                "full": full_emb,
                "img": img_emb,
                "skel": skel_emb,
            }

            # evaluate each target
            for t in targets:
                if t not in outs:
                    raise ValueError(f"unknown target {t}. available: {list(outs.keys())}")

                pred = outs[t]  # (N,D)
                pred = F.normalize(pred, dim=-1)

                # similarity to class candidates (N,C)
                sim = pred.float() @ class_emb.float().T

                update_metrics(metrics[t], sim, class_id_f, recall_ks)

    # save metrics + confusion
    summary_lines = []
    for t in targets:
        st = metrics[t]
        if st.total == 0:
            summary_lines.append(f"{t}: total=0 (no valid samples)")
            continue

        rec = {k: st.recall_hits[k] / st.total for k in recall_ks}
        summary_lines.append(f"{t}: total={st.total}, " + ", ".join([f"R@{k}={rec[k]:.4f}" for k in recall_ks]))

        # confusion outputs (per K)
        for k in recall_ks:
            k_dir = out_dir / f"top{k}"
            ensure_dir(k_dir)

            xlsx_path = k_dir / f"confusion_{t}.xlsx"
            png_path  = k_dir / f"confusion_{t}.png"

            save_confusion_excel(xlsx_path, st.confusion_by_k[k], sentences)
            save_confusion_png(png_path, st.confusion_by_k[k], normalize="true")


        # save per-target metrics yaml
        out_metrics = {
            "target": t,
            "total": st.total,
            "recall": {f"R@{k}": float(rec[k]) for k in recall_ks},
        }
        with (out_dir / f"metrics_{t}.yaml").open("w", encoding="utf-8") as f:
            yaml.safe_dump(out_metrics, f, allow_unicode=True, sort_keys=False)

    # print + save summary
    summary_text = "\n".join(summary_lines)
    print(summary_text)
    (out_dir / "summary.txt").write_text(summary_text + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
